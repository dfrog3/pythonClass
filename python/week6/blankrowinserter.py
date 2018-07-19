import sys

def blankRow(startRow, endRow, spreadsheet):
    import openpyxl
    from openpyxl.utils import get_column_letter, column_index_from_string


    toBlank = endRow - startRow

    wbo = openpyxl.load_workbook(spreadsheet)

    sheeto = wbo[wbo.sheetnames[0]]
    wbo.create_sheet(index=1, title= wbo.sheetnames[0])
    sheetn = wbo[wbo.sheetnames[1]]
    switchOn = False

    for row in range(1, sheeto.max_row + 1):
        for column in range(1, sheeto.max_column + 1):
            toCopyCell = (get_column_letter(column) + str(row))
            if row == startRow or switchOn:
                switchOn = True
                workingcell = (get_column_letter(column) + str(row + toBlank))
            else:
                workingcell = (get_column_letter(column) + str(row))

            # print("copy",toCopyCell)
            # print("paste",workingcell)
            sheetn[workingcell] = None
            # print("erase sheetn", workingcell)
            sheetn[workingcell] = sheeto[toCopyCell].value
            # print("doing copy")
            # print("sheeto",sheeto[toCopyCell].value)
    wbo.remove(sheeto)

    wbo.save(spreadsheet)
blankRow(int(sys.argv[1]),int(sys.argv[2]),sys.argv[3])
