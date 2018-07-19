import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string


wb = openpyxl.load_workbook("example.xlsx")

sheet = wb[wb.sheetnames[0]]
hi ="e1"
print(0, sheet[hi].value)
print(type(wb))
print(1,sheet[hi].value)
sheet[hi] = 6
print(2,sheet[hi].value)
print(3,sheet.cell(row = 1, column = 1).coordinate)
print(4,sheet)

for i in range(1, 8, 2):
    print(5,i, sheet.cell(row=i, column=2).value)


print(6,sheet.max_row, sheet.max_column)
print(7,get_column_letter(1),1)
print(8,column_index_from_string('A'),"A")