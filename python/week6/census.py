import openpyxl

wb = openpyxl.load_workbook("censuspopdata.xlsx")
sheet = wb[wb.sheetnames[0]]
bigBook = {}

for row in range(2, sheet.max_row+1):
    stateCell = "B" + str(row)
    countyCell = "C" + str(row)
    populationCell = "D" + str(row)


    state = sheet[stateCell].value
    county = sheet[countyCell].value
    population = sheet[populationCell].value


    bigBook.setdefault(state, {})
    bigBook[state].setdefault(county, {'tracts': 0, 'pop': 0})


    bigBook[state][county]["pop"] = bigBook[state][county]["pop"] + population
    bigBook[state][county]["tracts"] += 1









print(bigBook["AL"]["Autauga"]["tracts"])