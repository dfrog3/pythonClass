import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string


toInvertBook = openpyxl.load_workbook("switch.xlsx")
InvertedBook = openpyxl.load_workbook("inverted.xlsx")

sheetToInvert = toInvertBook[toInvertBook.sheetnames[0]]
sheetI = InvertedBook[InvertedBook.sheetnames[0]]

for row in range(1,sheetToInvert.max_row+1):
    for column in range(1,sheetToInvert.max_column + 1):
        to = (get_column_letter(row)+str(column))
        From = (get_column_letter(column)+str(row))
        sheetI[to] = sheetToInvert[From].value


InvertedBook.save("switch.xlsx")

