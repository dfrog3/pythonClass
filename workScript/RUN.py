import sys
import os
import openpyxl
from openpyxl.utils import get_column_letter
import PyPDF2
#required dependencies

def main(date, workBook):
    #"Graders Daily Schedule 2018_6(Jun).xlsx"
    scheduleBook = openpyxl.load_workbook(workBook)#gets workbook
    sheetNames = scheduleBook.sheetnames
    TEX = open("latexFiles/scheduleBlank.tex","r")


    sheetIndex = 0
    for sheet in sheetNames:
        if sheet == date:
            break
        else:
            sheetIndex += 1

    schedule = scheduleBook[scheduleBook.sheetnames[sheetIndex]]#gets sheet


    names = []
    for i in range(3,schedule.max_row+1):
        oneNameCell = ("e"+str(i))
        oneName = schedule[oneNameCell].value
        if oneName != None and schedule[("c"+str(i))].value != "休み":

            names.append(oneName)
            names.append(i)

    texFileNameList = []
    for i in range(0,len(names),2):
        listo =[]
        listo.append(names[i + 1])  # index 0 is row number

        listo.append(names[i])  # index 1 is name
        fileNameFile = listo[1].split(" ")

        newTexFileName = "output/" + fileNameFile[0]+fileNameFile[1] + ".tex"  # makes a new tex file named after child

        listo.append(newTexFileName)#index 2 is file name
        texFileNameList.append(listo)


    possibleTime = [1,1,1,1,1,1,1,1,"14:30",1,1,"15:00",1,1,"15:30",1,1,"16:00",1,1,"16:30",1,1,"17:00",1,1,"17:30",1,1,"18:00",1,1,"18:30",1,1,"19:00",1,1,"19:30"]
    #manages times slots. 1's fill out unused slots
    for oneList in texFileNameList:
        oneListIndex = 4
        lessonStart = False
        oneList.append(schedule[("g"+str(oneList[0]))].value)
        for i in range(8,37,3):
            cell = (get_column_letter(i)+str(oneList[0]))
            thisCellValue = schedule[cell].value
            if thisCellValue == None:
                thisCellValue = " "



            if thisCellValue == " " and not lessonStart:
                continue
            elif thisCellValue == "end":
                break
            elif lessonStart:
                oneList.append((possibleTime[i] + "-" +possibleTime[(i+3)] + " "+thisCellValue))
            elif thisCellValue =="Start":
                lessonStart = True
                continue
            else:
                lessonStart = True
                oneList.append((possibleTime[i] + "-" +possibleTime[(i+3)] + " "+thisCellValue))
                continue
    redLines = TEX.readlines()



    for oneList in texFileNameList:

        newFile = open(oneList[2],"w")
        stop = True


        for line in redLines:#adds information to the form. Each elif is one part of the form
            try:
                if "name" in line and stop:
                    newFile.write("﻿	"+oneList[1])
                elif "time" in line and stop:
                    newFile.write("﻿	"+oneList[3])
                elif "Date" in line and stop:
                    newFile.write(date)
                elif "slota" in line and stop:
                    if oneList[4] != None:
                        newFile.write("﻿	"+oneList[4])
                    else:
                        newFile.write("")
                elif "slotb" in line and stop:
                    if oneList[5] != None:
                        newFile.write("﻿	"+oneList[5])
                    else:
                        newFile.write("")
                elif "slotc" in line and stop:
                    if oneList[6] != None:
                        newFile.write("﻿	"+oneList[6])
                    else:
                        newFile.write("")
                elif "slotd" in line and stop:
                    if oneList[7] != None:
                        newFile.write("﻿	"+oneList[7])
                    else:
                        newFile.write("")
                elif "slote" in line and stop:
                    if oneList[8] != None:
                        newFile.write("﻿	"+oneList[8])
                    else:
                        newFile.write("")
                elif "slotf" in line and stop:
                    if oneList[9] != None:
                        newFile.write("﻿	"+oneList[9])
                    else:
                        newFile.write("")

                elif "slotg" in line and stop:
                    if oneList[10] != None:
                        newFile.write("﻿	"+oneList[10])
                    else:
                        newFile.write("")
                else:
                    newFile.write(line)
            except:
                newFile.write(r"﻿\end{itemForBlock}")
                newFile.write(r"﻿\begin{center}")
                newFile.write(r"﻿	\LARGE")
                newFile.write(r"﻿	Notes")
                newFile.write(r"﻿﻿\end{center}")
                newFile.write(r"﻿\end{document}")

                break


        newFile.close()
        fileNameFile = oneList[1].split(" ")

        x = ("rm "+fileNameFile[0]+fileNameFile[1] + ".log")
        y = ("rm "+fileNameFile[0]+fileNameFile[1] + ".aux")
        z = ("rm "+oneList[2])
        toMove = ("mv "+fileNameFile[0]+fileNameFile[1]+".pdf ./output/"+fileNameFile[0]+fileNameFile[1]+".pdf")
        toRun =("pdflatex ./" + oneList[2])
        os.system(toRun)
        os.system(toMove)

        os.system(z)
        os.system(x)
        os.system(y)







    pdfOutputFile = open('Print this four to a page.pdf', 'wb')
    pdfWriter = PyPDF2.PdfFileWriter()
    for oneList in texFileNameList:

        fileNameFile = oneList[1].split(" ")

        x = ("output/"+fileNameFile[0]+fileNameFile[1]+".pdf")





        try:

            pageToAdd = open(x,"rb")
            oneList.append(pageToAdd)
            pageToAddReader = PyPDF2.PdfFileReader(pageToAdd)
            onePage = pageToAddReader.getPage(0)
            pdfWriter.addPage(onePage)


        except:
            print("better define that varible on line 173")

    pdfWriter.write(pdfOutputFile)
    pdfOutputFile.close()
    for oneList in texFileNameList:
        x = (len(oneList)-1)
        oneList[x].close()





































#testing
#main("Thu 12", "7.10 mejiro (1).xlsx")
main(sys.argv[1],sys.argv[2])
